from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import load_workbook
from datetime import datetime
from datetime import date
import calendar
import json


def getPeople() -> object:
    import json
    f = open('data.json',)
    data = json.load(f)
    my_temp_arr = []
    for i in data['people']:
        my_temp_arr.append(i['name_surname'])
    return my_temp_arr

mydays = (calendar.day_name)

#VARIABLES
pageNumber = 6
rowNumber = 10
people = getPeople()

wb = Workbook()
ws = wb.active

# KISILERI YAZDIRMA ISLEME
for i, j in zip(range(len(people)), range(1, len(people) * 2 + 1)):
    ws.cell(row=2, column=j * 2, value=people[i])

# HEDEFLENEN VE SONUC YAZDIRMA ISLEMI
for i in range(2, len(people) * 2 + 1, 2):
    ws.cell(row=3, column=i, value='SS_1')
    ws.cell(row=3, column=i + 1, value='SS_2')

# SOL USTTEN ASAGI SAYILARI YAZDIRMA
for i in range(4, rowNumber + 4):
    ws.cell(row=i, column=1, value=i-3)

for i in range(2, len(people) * 2 + 1, 2):
    ws.merge_cells(start_row=2, start_column=i, end_row=2, end_column=i + 1)

# NEW PAGE ACMA

target = wb.copy_worksheet(wb.active)
target.title = "hey.xlsx"

wb.save('template.xlsx')


alignment=Alignment(horizontal='center', vertical='center', wrap_text=False, shrink_to_fit=False)
col_e = ws.column_dimensions['C']
col_e.alignment = alignment


now = datetime.now()
today = now.strftime("%d") + "_" + now.strftime("%m")+ "_" + now.strftime("%Y")
dayName = now.strftime("%A")
